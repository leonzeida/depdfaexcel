#!/usr/bin/env python3
"""Extrae la tabla de items de un Pedido de Cotizacion F.41
(Gobierno de la Provincia de Formosa) y la vuelca a un Excel.

Uso:
    python3 f41_a_excel.py archivo1.pdf archivo2.pdf ...
    python3 f41_a_excel.py carpeta_con_pdfs/

Por cada PDF de entrada se genera un .xlsx con el mismo nombre, en la
misma carpeta que el PDF.
"""

import re
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.filters import AutoFilter, FilterColumn
from openpyxl.worksheet.table import Table, TableStyleInfo

# Coordenadas (en puntos) de las columnas de la tabla "DETALLE DE ITEMS"
# del formulario F.41. Son fijas porque el formulario es una plantilla
# oficial con posiciones constantes.
COL_RG_MAX = 95
COL_CODIGO_MAX = 168
COL_DESC_MAX = 384
COL_CANT_MAX = 420
LEFT_BORDER_X = 70.9
LEFT_BORDER_TOL = 2

# Sólo se conservan los items cuyo código empieza con alguno de estos prefijos.
PREFIJOS_CODIGO = ("4.01.008", "4.01.018")

# Datos del encabezado del F.41 (página 1), usados para armar la Nota de
# Pedido y para nombrar los archivos de salida.
RE_EXPEDIENTE = re.compile(r"Expediente:\s*[A-Z]?-?(\d+/\d+)")
RE_CONTRATACION = re.compile(r"CONTRATACI[ÓO]N DIRECTA N[º°]:\s*(\S+)")
RE_APERTURA = re.compile(r"APERTURA:\s*(\S+)")
RE_HORA = re.compile(r"HORA:\s*([\d:]+)")


def bandas_de_filas(page):
    """Detecta las bandas verticales (top, bottom) de cada fila de la
    tabla a partir de los segmentos del borde izquierdo."""
    segmentos = [
        l
        for l in page.lines
        if abs(l["x0"] - l["x1"]) < 0.5 and abs(l["x0"] - LEFT_BORDER_X) < LEFT_BORDER_TOL
    ]
    segmentos.sort(key=lambda l: l["top"])
    return [(s["top"], s["bottom"]) for s in segmentos]


def extraer_filas_pagina(page):
    palabras = page.extract_words()
    filas = []
    for top, bottom in bandas_de_filas(page):
        en_banda = [w for w in palabras if top - 0.5 <= w["top"] < bottom - 0.5]
        if not en_banda:
            continue

        rg_palabras = [w for w in en_banda if w["x0"] < COL_RG_MAX]
        codigo_palabras = [w for w in en_banda if COL_RG_MAX <= w["x0"] < COL_CODIGO_MAX]
        # La columna "Cantidad" está alineada a la derecha, así que un
        # número largo (ej. "152064") puede empezar (x0) antes del límite
        # de la columna de Descripción aunque termine (x1) bien adentro de
        # la columna de Cantidad. Por eso esta columna se distingue por su
        # borde derecho (x1) y no por el izquierdo (x0).
        desc_palabras = [
            w
            for w in en_banda
            if COL_CODIGO_MAX <= w["x0"] < COL_DESC_MAX and w["x1"] <= COL_DESC_MAX
        ]
        cant_palabras = [
            w
            for w in en_banda
            if w["x0"] >= COL_CODIGO_MAX and COL_DESC_MAX < w["x1"] <= COL_CANT_MAX
        ]

        rg_texto = "".join(w["text"] for w in rg_palabras)
        if not rg_texto.isdigit():
            # No es una fila de item real (encabezado de columnas, fila
            # de "Total", etc.)
            continue

        codigo = "".join(w["text"] for w in codigo_palabras)
        desc_palabras.sort(key=lambda w: (round(w["top"], 1), w["x0"]))
        descripcion = " ".join(w["text"] for w in desc_palabras)
        cant_texto = "".join(w["text"] for w in cant_palabras)
        try:
            cantidad = int(cant_texto)
        except ValueError:
            cantidad = cant_texto

        filas.append(
            {
                "rg": int(rg_texto),
                "codigo": codigo,
                "descripcion": descripcion,
                "cantidad": cantidad,
            }
        )
    return filas


def extraer_pdf(pdf_path: Path):
    with pdfplumber.open(pdf_path) as pdf:
        filas = []
        for page in pdf.pages:
            filas.extend(extraer_filas_pagina(page))
    return [f for f in filas if f["codigo"].startswith(PREFIJOS_CODIGO)]


def extraer_encabezado(pdf_path: Path):
    with pdfplumber.open(pdf_path) as pdf:
        texto = pdf.pages[0].extract_text() or ""

    def buscar(patron, nombre):
        m = patron.search(texto)
        if not m:
            raise ValueError(f"No se encontró '{nombre}' en el encabezado del PDF.")
        return m.group(1)

    return {
        "expediente": buscar(RE_EXPEDIENTE, "N°Expediente"),
        "contratacion": buscar(RE_CONTRATACION, "Contratación Directa"),
        "apertura": buscar(RE_APERTURA, "Apertura"),
        "hora": buscar(RE_HORA, "Hora"),
    }


def expediente_slug(expediente: str) -> str:
    return expediente.replace("/", "-")


def nombre_archivo_notas_pedido(encabezado: dict) -> str:
    expediente_u = "U-" + expediente_slug(encabezado["expediente"])
    dia, mes, anio = encabezado["apertura"].split("/")
    fecha = f"{dia}-{mes}-{anio[-2:]}"
    hora = encabezado["hora"].split(":")[0] + "Hs"
    return f"EXPT  N° {expediente_u}   {fecha} {hora}"


ZONA_HORARIA_ARGENTINA = ZoneInfo("America/Argentina/Buenos_Aires")


def linea_de_pedido(encabezado: dict) -> str:
    hora_fmt = encabezado["hora"].replace(":", ",") + " HS"
    return (
        f"Nota de Pedido {encabezado['contratacion']} {encabezado['apertura']} "
        f"{hora_fmt} Expt {encabezado['expediente']}"
    )


# Formatos de moneda usados en las planillas de Zeid Medical (tal cual los
# archivos de ejemplo que armó el cliente).
FMT_MONEDA = r'_ "$"\ * #,##0.00_ ;_ "$"\ * \-#,##0.00_ ;_ "$"\ * "-"??_ ;_ @_ '
FMT_MONEDA_ARS = r'_-[$$-2C0A]\ * #,##0.00_-;\-[$$-2C0A]\ * #,##0.00_-;_-[$$-2C0A]\ * "-"??_-;_-@_-'


def escribir_notas_pedido(filas: list, encabezado: dict, salida: Path):
    wb = Workbook()
    ws = wb.active
    slug = expediente_slug(encabezado["expediente"])
    ws.title = slug[:31]

    hoy = datetime.now(ZONA_HORARIA_ARGENTINA).strftime("%d/%m/%y")
    linea_pedido = linea_de_pedido(encabezado)

    fuente_titulo = Font(name="ITC Zapf Chancery", bold=True, size=36)
    fuente_membrete = Font(name="Book Antiqua", bold=True, size=10)
    fuente_pedido = Font(name="Courier New", bold=True, size=16, color="000000")
    fuente_encabezado_tabla = Font(name="Arial", bold=True, size=10, color="000000")
    fuente_nota = Font(name="Arial", size=11)
    fuente_total = Font(name="Arial", bold=True, size=11)

    centrado = Alignment(horizontal="center")
    centrado_wrap = Alignment(horizontal="center", wrap_text=True)
    centrado_medio = Alignment(horizontal="center", vertical="center")
    relleno_encabezado = PatternFill("solid", fgColor="D4EA6B")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Zeid Medical S.R.L"
    ws["A1"].font = fuente_titulo
    ws["A1"].alignment = centrado
    ws.row_dimensions[1].height = 39.75

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "Av.25 de Mayo 1085-  Formosa"
        + " " * 90
        + hoy
    )
    ws["A2"].font = fuente_membrete
    ws["A2"].alignment = centrado_wrap
    ws.row_dimensions[2].height = 15

    ws.merge_cells("A3:E3")
    ws["A3"] = (
        "        Cuit: 30-70849078-0                                                                                                         "
        "IVA responsable Inscripto"
    )
    ws["A3"].font = fuente_membrete
    ws["A3"].alignment = centrado

    ws.merge_cells("A4:E4")
    ws["A4"] = (
        "  Ingresos Brutos: 30-70849078-0                                                                                    "
        "Inicio de Actividad:07/2009"
    )
    ws["A4"].font = fuente_membrete
    ws["A4"].alignment = centrado

    ws.merge_cells("A5:E5")
    ws["A5"] = linea_pedido
    ws["A5"].font = fuente_pedido
    ws["A5"].alignment = centrado
    ws.row_dimensions[5].height = 21.4
    # El borde y el relleno se aplican a las 5 celdas del rango combinado:
    # en una celda combinada, Excel toma el borde derecho de la última
    # columna (E) y no de la primera (A), así que si sólo se lo ponemos a
    # A5 el lado derecho queda sin cerrar.
    for c in range(1, 6):
        celda = ws.cell(row=5, column=c)
        celda.fill = relleno_encabezado
        celda.border = Border(
            top=Side(style="medium"),
            bottom=Side(style="medium"),
            left=Side(style="medium") if c == 1 else None,
            right=Side(style="medium") if c == 5 else None,
        )

    # Encabezado de la tabla en una única fila (antes "Reng" e "Importe
    # total" ocupaban 2 filas combinadas mientras el resto sólo 1, dejando
    # un hueco sin estilo debajo de Descripción/Cantidad/Precio).
    encabezados_tabla = ["Renglones", "Descripción", "Cantidad", "Precio", "Importe total"]
    for col, titulo in enumerate(encabezados_tabla, start=1):
        c = ws.cell(row=6, column=col, value=titulo)
        c.alignment = centrado_wrap if col == 2 else centrado_medio
        c.fill = relleno_encabezado
        c.font = fuente_encabezado_tabla
    ws.row_dimensions[6].height = 18

    fila = 7
    primera_fila_datos = fila
    for item in filas:
        c_reng = ws.cell(row=fila, column=1, value=item["rg"])
        c_reng.alignment = centrado
        ws.cell(row=fila, column=2, value=item["descripcion"])
        c_cant = ws.cell(row=fila, column=3, value=item["cantidad"])
        c_cant.alignment = centrado
        c_precio = ws.cell(row=fila, column=4)
        c_precio.number_format = FMT_MONEDA
        c_importe = ws.cell(row=fila, column=5, value=f"=C{fila}*D{fila}")
        c_importe.number_format = FMT_MONEDA
        fila += 1
    ultima_fila_datos = fila - 1

    borde_negro = Side(style="thin", color="000000")
    primera_fila_tabla = 6
    for r in range(primera_fila_tabla, ultima_fila_datos + 1):
        for c in range(1, 6):
            if r not in (primera_fila_tabla, ultima_fila_datos) and c not in (1, 5):
                continue
            ws.cell(row=r, column=c).border = Border(
                top=borde_negro if r == primera_fila_tabla else None,
                bottom=borde_negro if r == ultima_fila_datos else None,
                left=borde_negro if c == 1 else None,
                right=borde_negro if c == 5 else None,
            )

    tabla = Table(
        displayName=f"TablaNotaPedido_{slug.replace('-', '_')}",
        ref=f"A6:E{ultima_fila_datos}",
    )
    # openpyxl vuelve a crear un autoFilter vacío al guardar si la tabla
    # tiene encabezado, así que no alcanza con ponerlo en None: hay que
    # ocultar el botón de cada columna explícitamente (hiddenButton).
    tabla.autoFilter = AutoFilter(
        ref=tabla.ref,
        filterColumn=[FilterColumn(colId=i, hiddenButton=True) for i in range(5)],
    )
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showRowStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
    )
    ws.add_table(tabla)

    c_nota = ws.cell(
        row=fila,
        column=2,
        value="NOTA : LOS PRODUCTOS COTIZADOS TIENEN QUE TENER UN VENCIMIENTO DE 12 MESES O MAYOR.",
    )
    c_nota.font = fuente_nota
    c_gran_total_label = ws.cell(row=fila, column=4, value="GRAN TOTAL")
    c_gran_total_label.font = fuente_total
    c_gran_total = ws.cell(row=fila, column=5, value=f"=SUM(E{primera_fila_datos}:E{ultima_fila_datos})")
    c_gran_total.font = fuente_total
    c_gran_total.number_format = FMT_MONEDA

    anchos = {"A": 11.4, "B": 98.27, "C": 10.93, "D": 13.93, "E": 13.6}
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    wb.save(salida)


def escribir_planilla_trabajo(filas: list, encabezado: dict, salida: Path):
    wb = Workbook()
    ws = wb.active
    slug = expediente_slug(encabezado["expediente"])
    ws.title = slug[:31]

    fuente_pedido = Font(name="Courier New", bold=True, size=16, color="000000")
    fuente_encabezado = Font(bold=True, size=12, color="000000")
    fuente_encabezado_chico = Font(bold=True, size=11, color="000000")

    centrado = Alignment(horizontal="center")
    izquierda = Alignment(horizontal="left")
    relleno_encabezado = PatternFill("solid", fgColor="D4EA6B")

    ws.merge_cells("A1:J1")
    ws["A1"] = linea_de_pedido(encabezado)
    ws["A1"].font = fuente_pedido
    ws["A1"].alignment = centrado
    ws.row_dimensions[1].height = 21.4
    # Mismo criterio que la fila 5 de Nota de Pedido: el relleno y el
    # borde se aplican a las 10 celdas del rango combinado, porque en una
    # celda combinada Excel toma el borde derecho de la última columna
    # (J) y no de la primera (A).
    for c in range(1, 11):
        celda = ws.cell(row=1, column=c)
        celda.fill = relleno_encabezado
        celda.border = Border(
            top=Side(style="medium"),
            bottom=Side(style="medium"),
            left=Side(style="medium") if c == 1 else None,
            right=Side(style="medium") if c == 10 else None,
        )

    encabezados_tabla = [
        ("A", "Renglon", centrado, fuente_encabezado, None),
        ("B", "Codigos", centrado, fuente_encabezado, None),
        ("C", "Descripcion", centrado, fuente_encabezado, None),
        ("D", "Cantidad", izquierda, fuente_encabezado, None),
        ("E", "Costo", centrado, fuente_encabezado_chico, FMT_MONEDA),
        ("F", "Total costo", centrado, fuente_encabezado_chico, None),
        ("G", "Precio ref.", centrado, fuente_encabezado_chico, FMT_MONEDA),
        ("H", "Pcio Vta", centrado, fuente_encabezado_chico, None),
        ("I", "Total de Vta", centrado, fuente_encabezado_chico, None),
        ("J", "%", centrado, fuente_encabezado_chico, None),
    ]
    for letra, titulo, alineacion, fuente, numfmt in encabezados_tabla:
        c = ws[f"{letra}2"]
        c.value = titulo
        c.font = fuente
        c.alignment = alineacion
        c.fill = relleno_encabezado
        if numfmt:
            c.number_format = numfmt
    ws.row_dimensions[2].height = 15.75

    fila = 3
    primera_fila_datos = fila
    for item in filas:
        ws.cell(row=fila, column=1, value=item["rg"]).alignment = centrado
        ws.cell(row=fila, column=2, value=item["codigo"]).alignment = centrado
        ws.cell(row=fila, column=3, value=item["descripcion"])
        ws.cell(row=fila, column=4, value=item["cantidad"]).alignment = izquierda
        c_total_costo = ws.cell(row=fila, column=6, value=f"=D{fila}*E{fila}")
        c_total_costo.alignment = centrado
        c_total_costo.number_format = FMT_MONEDA
        c_pcio_vta = ws.cell(row=fila, column=8, value=f"=E{fila}*1.5")
        c_pcio_vta.alignment = centrado
        c_pcio_vta.number_format = FMT_MONEDA_ARS
        c_total_vta = ws.cell(row=fila, column=9, value=f"=H{fila}*D{fila}")
        c_total_vta.alignment = centrado
        c_total_vta.number_format = FMT_MONEDA_ARS
        c_pct = ws.cell(row=fila, column=10, value=f"=H{fila}/E{fila}-1")
        c_pct.alignment = centrado
        c_pct.number_format = "0%"
        fila += 1
    ultima_fila_datos = fila - 1

    tabla = Table(
        displayName=f"TablaPlanilla_{slug.replace('-', '_')}",
        ref=f"A2:J{ultima_fila_datos}",
    )
    # Mismo fix que en Nota de Pedido: hay que ocultar el botón de cada
    # columna explícitamente porque openpyxl recrea un autoFilter vacío
    # al guardar si la tabla tiene fila de encabezado.
    tabla.autoFilter = AutoFilter(
        ref=tabla.ref,
        filterColumn=[FilterColumn(colId=i, hiddenButton=True) for i in range(10)],
    )
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showRowStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
    )
    ws.add_table(tabla)

    c_total_costo_general = ws.cell(row=fila, column=6, value=f"=SUM(F{primera_fila_datos}:F{ultima_fila_datos})")
    c_total_costo_general.number_format = FMT_MONEDA
    c_total_vta_general = ws.cell(row=fila, column=9, value=f"=SUM(I{primera_fila_datos}:I{ultima_fila_datos})")
    c_total_vta_general.number_format = FMT_MONEDA_ARS

    anchos = {
        "A": 12, "B": 16, "C": 100, "D": 15.4, "E": 11.4,
        "F": 15.2, "G": 11.8, "H": 11.33, "I": 16.46, "J": 10.66,
    }
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    wb.save(salida)


def procesar(pdf_path: Path):
    print(f"Procesando: {pdf_path.name}")
    filas = extraer_pdf(pdf_path)
    if not filas:
        print(f"  ADVERTENCIA: no se detectaron items en {pdf_path.name}")
    encabezado = extraer_encabezado(pdf_path)
    slug = expediente_slug(encabezado["expediente"])

    salida_np = pdf_path.parent / f"{nombre_archivo_notas_pedido(encabezado)}.xlsx"
    escribir_notas_pedido(filas, encabezado, salida_np)
    print(f"  -> {salida_np.name} ({len(filas)} items)")

    salida_pt = pdf_path.parent / f"Planilla de Trabajo {slug}.xlsx"
    escribir_planilla_trabajo(filas, encabezado, salida_pt)
    print(f"  -> {salida_pt.name} ({len(filas)} items)")


def main(argv):
    if not argv:
        print(__doc__)
        return 1

    pdfs = []
    for arg in argv:
        p = Path(arg)
        if p.is_dir():
            pdfs.extend(sorted(p.glob("*.pdf")))
        elif p.is_file():
            pdfs.append(p)
        else:
            print(f"AVISO: no existe el archivo/carpeta: {arg}")

    if not pdfs:
        print("No se encontraron archivos PDF para procesar.")
        return 1

    errores = 0
    for pdf_path in pdfs:
        try:
            procesar(pdf_path)
        except Exception as exc:  # noqa: BLE001 - queremos seguir con el resto
            errores += 1
            print(f"  ERROR procesando {pdf_path.name}: {exc}")

    return 1 if errores else 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
